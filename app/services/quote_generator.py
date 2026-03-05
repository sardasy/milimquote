import io
from datetime import timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models.models import Quote


def _border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style() -> tuple[Font, PatternFill]:
    font = Font(bold=True, color="FFFFFF", size=11)
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    return font, fill


def generate_quote_excel(quote: Quote) -> bytes:
    """Quote ORM 객체를 받아 Excel 견적서를 bytes로 반환한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    # 열 너비
    col_widths = {"A": 6, "B": 38, "C": 10, "D": 18, "E": 20, "F": 18}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── 제목 ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "견  적  서"
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45

    # ── 기본 정보 ─────────────────────────────────────────────────────────────
    rate = quote.exchange_rate_ref.exchange_rate
    valid_until = (quote.created_at + timedelta(days=30)).strftime("%Y-%m-%d")

    info_rows = [
        ("견적번호", quote.quote_number, "견적일", quote.created_at.strftime("%Y-%m-%d")),
        ("고객사", quote.customer.name, "유효기간", valid_until),
        ("담당자", quote.customer.contact_name or "-", "납기", f"{quote.delivery_days}일"),
        ("적용환율", f"1 USD = {rate:,.2f} KRW", "", ""),
    ]

    label_font = Font(bold=True)
    row = 2
    for left_label, left_val, right_label, right_val in info_rows:
        ws.cell(row=row, column=1, value=left_label).font = label_font
        ws.cell(row=row, column=2, value=left_val)
        ws.cell(row=row, column=4, value=right_label).font = label_font
        ws.cell(row=row, column=5, value=right_val)
        row += 1

    row += 1  # 빈 줄

    # ── 품목 테이블 헤더 ──────────────────────────────────────────────────────
    header_font, header_fill = _header_style()
    headers = ["No.", "제품명", "수량", "단가(원)", "금액(원)", "비고"]
    header_row = row
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
    ws.row_dimensions[header_row].height = 22
    row += 1

    # ── 품목 행 ───────────────────────────────────────────────────────────────
    total_amount = 0
    for idx, item in enumerate(quote.items, 1):
        line_total = item.unit_price_krw * item.quantity
        total_amount += line_total

        row_data = [idx, item.product.name, item.quantity, item.unit_price_krw, line_total, ""]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _border()
            if col == 2:
                cell.alignment = Alignment(horizontal="left")
            elif col in (3,):
                cell.alignment = Alignment(horizontal="center")
            elif col in (4, 5):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
        row += 1

    # ── 합계 행 ───────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    total_label = ws.cell(row=row, column=1, value="합  계")
    total_label.font = Font(bold=True)
    total_label.alignment = Alignment(horizontal="right")
    total_label.border = _border()

    total_cell = ws.cell(row=row, column=5, value=total_amount)
    total_cell.font = Font(bold=True)
    total_cell.number_format = "#,##0"
    total_cell.border = _border()
    total_cell.alignment = Alignment(horizontal="right")

    ws.cell(row=row, column=6).border = _border()
    row += 2

    # ── 비고 ─────────────────────────────────────────────────────────────────
    if quote.notes:
        ws.cell(row=row, column=1, value="비고").font = Font(bold=True)
        ws.cell(row=row, column=2, value=quote.notes)
        row += 1

    ws.cell(row=row, column=1, value="※ 상기 금액은 부가세 별도입니다.").font = Font(
        italic=True, color="666666", size=9
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
